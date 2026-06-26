#!/usr/bin/env python3
"""
generate_dashboard.py  <input.xlsx>  <output.html>

Reads a calculation-log Excel file (produced by parsecalclogdetail.py),
and writes a self-contained interactive HTML dashboard.

Supports 1–N runs (Excel sheets). Comparison defaults to first vs last.
HTML structure lives in dashboard_template.html alongside this script.
"""

import json
import os
import re
import sqlite3
import sys
import zipfile
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    print("openpyxl not found. Run: pip install openpyxl --break-system-packages")
    sys.exit(1)

SCRIPT_DIR    = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_PATH = os.path.join(SCRIPT_DIR, 'dashboard_template.html')
DATA_MARKER   = '__DATA_PLACEHOLDER__'
BATCH_RE      = re.compile(r'Batch "([^"]+)"')
GENERIC_TRANSFORMS = {"Plan Calculation", "Crediting Component"}

# Queries larger than this are truncated in the dashboard (they're unreadable anyway).
MAX_QUERY_CHARS = 50_000
TEXT_TABLES = {
    "queries": "query_text",
    "details": "detail_text",
}


# ── Excel / SQLite loading ────────────────────────────────────────────────────

def resolve_text(text_raw, xlsx_dir, db_con, table_name):
    """
    Return (display_text, db_id) for a cell value, or (None, None).

    Handles three formats:
      - int / float  → SQLite row ID (current format)
      - string int   → SQLite row ID stored as text
      - "*.txt" path → legacy flat-file format
    """
    if text_raw is None:
        return None, None   # (display_text, db_id_for_export)

    text_column = TEXT_TABLES[table_name]

    if isinstance(text_raw, (int, float)):
        db_id = int(text_raw)
        row = db_con.execute(f"SELECT {text_column} FROM {table_name} WHERE id = ?", (db_id,)).fetchone() if db_con else None
        text = row[0] if row else f"[Query id={db_id} not found in DB]"
        return _truncate(text), db_id

    q = str(text_raw).strip()

    if q.lstrip("-").isdigit() and db_con:
        db_id = int(q)
        row = db_con.execute(f"SELECT {text_column} FROM {table_name} WHERE id = ?", (db_id,)).fetchone()
        text = row[0] if row else f"[Query id={db_id} not found in DB]"
        return _truncate(text), db_id

    if q.endswith(".txt") and xlsx_dir:
        try:
            with open(os.path.join(xlsx_dir, q), encoding="utf-8") as fh:
                text = fh.read()
            return _truncate(text), None   # no DB id for legacy flat files
        except OSError:
            msg = f"[Query file not found: {q}]"
            return msg, None

    return _truncate(q), None


def resolve_query(query_raw, xlsx_dir, db_con):
    return resolve_text(query_raw, xlsx_dir, db_con, "queries")


def resolve_detail(detail_raw, xlsx_dir, db_con):
    return resolve_text(detail_raw, xlsx_dir, db_con, "details")


def _truncate(text):
    if text and len(text) > MAX_QUERY_CHARS:
        return text[:MAX_QUERY_CHARS] + f"\n\n… [truncated — {len(text):,} chars total, showing first {MAX_QUERY_CHARS:,}]"
    return text


def load_sheet(ws, xlsx_dir, db_con):
    """Read one Excel sheet into a list of row dicts, plus the trace ID and earliest start datetime."""
    rows     = []
    trace_id = None
    min_start_dt = None
    headers = {
        str(ws.cell(1, c).value): c
        for c in range(1, ws.max_column + 1)
        if ws.cell(1, c).value is not None
    }

    def cell(row_idx, names, fallback_col=None):
        for name in names:
            if name in headers:
                return ws.cell(row_idx, headers[name]).value
        if fallback_col is not None:
            return ws.cell(row_idx, fallback_col).value
        return None

    for r in range(2, ws.max_row + 1):
        transform = cell(r, ("transform",), 3)
        if transform is None:
            continue

        step       = cell(r, ("step",), 2)
        start_dt   = cell(r, ("start_time",), 4)
        duration   = cell(r, ("duration_in_seconds", "duration"), 6) or 0
        records    = cell(r, ("records_updated", "records"), 7)
        query_raw  = cell(r, ("query",), 8)
        detail_raw = cell(r, ("detail",))
        tid        = cell(r, ("trace_id",), 9)

        if tid and trace_id is None:
            trace_id = str(tid)

        if start_dt is not None and hasattr(start_dt, "isoformat"):
            if min_start_dt is None or start_dt < min_start_dt:
                min_start_dt = start_dt

        batch_match = BATCH_RE.search(str(step or ""))
        batch = batch_match.group(1) if batch_match else "Plan Calculations (untagged)"

        display_name = str(step or "") if str(transform) in GENERIC_TRANSFORMS and step else str(transform)
        _q = resolve_query(query_raw, xlsx_dir, db_con)
        _detail = resolve_detail(detail_raw, xlsx_dir, db_con)
        rows.append({
            "step":        str(step or ""),
            "transform":   str(transform),
            "display_name": display_name,
            "batch":       batch,
            "source_index": len(rows) + 1,
            "start_time":  start_dt.strftime("%H:%M:%S") if start_dt else None,
            "duration":    float(duration),
            "records":     int(records) if records is not None else None,
            "query":       _q[0],        # truncated display text
            "query_db_id": _q[1],        # SQLite row id (int) or None
            "detail":      _detail[0],
            "detail_db_id": _detail[1],
        })

    return rows, trace_id, min_start_dt


def open_db_if_exists(xlsx_abs):
    """Open the sibling .db file if present, otherwise return None."""
    base   = os.path.splitext(os.path.basename(xlsx_abs))[0]
    db_path = os.path.join(os.path.dirname(xlsx_abs), base + ".db")
    if os.path.exists(db_path):
        print(f"  Using query DB: {db_path}")
        return sqlite3.connect(db_path)
    return None


# ── Data assembly ─────────────────────────────────────────────────────────────

def batch_totals(rows):
    totals = defaultdict(float)
    for row in rows:
        totals[row["batch"]] += row["duration"]
    return {k: round(v, 3) for k, v in totals.items()}


def earliest_batch_starts(rows):
    starts = {}
    for row in rows:
        b, st = row["batch"], row["start_time"]
        if b not in starts and st:
            starts[b] = st
    return starts


def align_transforms(runs_raw):
    """
    Align per-transform rows across all runs by (step, transform, occurrence).
    Returns a list of transform dicts, each with per-run duration/records/start_time lists.
    """
    n = len(runs_raw)

    # Per-run lookup: (step, transform) → ordered list of rows
    run_dicts = []
    for run in runs_raw:
        d = defaultdict(list)
        for row in run["rows"]:
            d[(row["step"], row["transform"])].append(row)
        # When the same sub-step repeats (e.g. "for Period 78" × 8), the execution
        # order can differ across runs. Sort by duration so the shortest sub-step in
        # run A is matched with the shortest in run B, not the one at the same log
        # position — which would be wrong whenever the order changed.
        for rows in d.values():
            if len(rows) > 1:
                rows.sort(key=lambda r: r["duration"])
        run_dicts.append(d)

    # Collect all (step, transform, occurrence) keys in global order
    seen, ordered_keys = {}, []
    for run in runs_raw:
        counts = defaultdict(int)
        for row in run["rows"]:
            base = (row["step"], row["transform"])
            counts[base] += 1
            key = (*base, counts[base])
            if key not in seen:
                seen[key] = len(ordered_keys)
                ordered_keys.append(key)

    # Build aligned transforms
    cursors = [defaultdict(int) for _ in runs_raw]
    transforms = []
    for step_val, transform_val, _occ in ordered_keys:
        base     = (step_val, transform_val)
        row_data = []
        batch    = None

        for ri, (d, cursor) in enumerate(zip(run_dicts, cursors)):
            available = d[base]
            idx = cursor[base]
            if idx < len(available):
                row = available[idx]
                cursor[base] += 1
                batch = batch or row["batch"]
                row_data.append(row)
            else:
                row_data.append(None)

        if batch is None:
            m = BATCH_RE.search(step_val)
            batch = m.group(1) if m else "Plan Calculations (untagged)"

        display_name = next((r["display_name"] for r in row_data if r), transform_val)
        transforms.append({
            "batch":       batch,
            "step":        step_val,
            "transform":   display_name,
            "event":       transform_val,
            "sourceIndexes": [r["source_index"] if r else None for r in row_data],
            "start_times": [r["start_time"] if r else None for r in row_data],
            "durations":   [round(r["duration"], 3) if r else None for r in row_data],
            "records":     [r["records"] if r else None for r in row_data],
            "queries":     [r["query"] if r else None for r in row_data],
            "queryDbIds":  [r["query_db_id"] if r else None for r in row_data],
            "details":     [r["detail"] if r else None for r in row_data],
            "detailDbIds": [r["detail_db_id"] if r else None for r in row_data],
        })

    same_length = all(all(d is not None for d in t["durations"]) for t in transforms)
    return transforms, same_length


def has_same_step_sequence(runs_raw):
    """Return True when every run has the same parsed step/event order."""
    if len(runs_raw) <= 1:
        return True

    first = [(row["step"], row["transform"]) for row in runs_raw[0]["rows"]]
    return all(
        [(row["step"], row["transform"]) for row in run["rows"]] == first
        for run in runs_raw[1:]
    )


def compute_regressions(transforms, ia, ib):
    regs = []
    for t in transforms:
        da, db = t["durations"][ia], t["durations"][ib]
        if da and db:
            diff = round(db - da, 3)
            regs.append({
                "batch":     t["batch"],
                "transform": t["transform"],
                "da": da, "db": db,
                "diff": diff,
                "pct": round(diff / da * 100, 1),
            })
    return sorted(regs, key=lambda x: -x["diff"])[:25]


def build_data(xlsx_path):
    xlsx_abs = os.path.abspath(xlsx_path)
    xlsx_dir = os.path.dirname(xlsx_abs)
    wb       = openpyxl.load_workbook(xlsx_abs)

    if not wb.sheetnames:
        print("No sheets found in Excel file.")
        sys.exit(1)

    db_con = open_db_if_exists(xlsx_abs)

    # Load each sheet into a run dict
    runs_raw = []
    for name in wb.sheetnames:
        rows, trace_id, min_start_dt = load_sheet(wb[name], xlsx_dir, db_con)
        totals = batch_totals(rows)
        runs_raw.append({
            "label":         name,
            "total":         round(sum(r["duration"] for r in rows), 2),
            "steps":         len(rows),
            "rows":          rows,
            "batch_totals":  totals,
            "trace_id":      trace_id,
            "min_start_dt":  min_start_dt,
        })

    # Sort runs chronologically so default_a=first and default_b=last are meaningful.
    # Excel sheet order is insertion order (often alphabetical by filename), which
    # doesn't match run order when files are named by duration (e.g. 2H46M, 3H51M).
    # Runs with no parseable start time sort last (1) so they don't become the baseline.
    def _start_dt_key(run):
        dt = run.get("min_start_dt")
        if dt is None or not hasattr(dt, "isoformat"):
            return (1, "")
        return (0, dt.isoformat())

    runs_raw.sort(key=_start_dt_key)

    if db_con:
        db_con.close()

    n  = len(runs_raw)
    ia, ib = 0, n - 1

    # Batches sorted by peak duration descending
    peak = defaultdict(float)
    for run in runs_raw:
        for b, d in run["batch_totals"].items():
            peak[b] = max(peak[b], d)

    batch_order = {}
    for row in runs_raw[0]["rows"]:
        if row["batch"] not in batch_order:
            batch_order[row["batch"]] = len(batch_order)

    batches = [
        {
            "name":        b,
            "durations":   [run["batch_totals"].get(b, 0) for run in runs_raw],
            "start_times": [earliest_batch_starts(run["rows"]).get(b) for run in runs_raw],
            "orig_idx":    batch_order.get(b, 9999),
        }
        for b in sorted(peak, key=lambda b: -peak[b])
    ]

    transforms, same_length = align_transforms(runs_raw)
    same_sequence = has_same_step_sequence(runs_raw)
    regressions = compute_regressions(transforms, ia, ib) if ia != ib else []

    worst   = regressions[0] if regressions else None
    ta      = runs_raw[ia]["total"]
    tb      = runs_raw[ib]["total"] if n > 1 else None
    reg_pct = round((tb - ta) / ta * 100, 1) if tb else None

    # Print summary
    print(f"  Sheets ({n}): " + " / ".join(r["label"] for r in runs_raw))
    for r in runs_raw:
        print(f"    {r['label']}: {r['total']}s ({r['steps']} steps)")
    if reg_pct is not None:
        sign = "+" if reg_pct > 0 else ""
        print(f"  First vs Last Regression: {sign}{reg_pct}%")

    return {
        "runs":            [{"label": r["label"], "total": r["total"], "steps": r["steps"], "trace_id": r.get("trace_id")} for r in runs_raw],
        "default_a":       ia,
        "default_b":       ib,
        "batches":         batches,
        "transforms":      transforms,
        "regression_pct":  reg_pct,
        "worst_step_diff": round(worst["diff"], 1) if worst else None,
        "worst_step_name": worst["transform"] if worst else None,
        "same_length":     same_length,
        "same_sequence":   same_sequence,
        "match_mode":      "step + event + occurrence",
    }


# ── Entry point ───────────────────────────────────────────────────────────────

def _safe_filename(name):
    """Sanitise a transform name for use as a filename."""
    return re.sub(r'[\\/:*?"<>|]', '_', name).strip() or 'query'


def generate(xlsx_path, output_path):
    print(f"Reading: {xlsx_path}")
    data = build_data(xlsx_path)

    # ── Write per-transform .log files ────────────────────────────────────────
    # Each transform × run gets its own .log file in a _logs/ subfolder so the
    # export button can just trigger a plain anchor download — no file picker.
    xlsx_abs  = os.path.abspath(xlsx_path)
    db_path   = os.path.splitext(xlsx_abs)[0] + ".db"
    db_con    = sqlite3.connect(db_path) if os.path.exists(db_path) else None

    out_base  = os.path.splitext(os.path.abspath(output_path))[0]
    logs_dir  = out_base + "_logs"
    logs_name = os.path.basename(logs_dir)   # relative folder name for HTML hrefs
    os.makedirs(logs_dir, exist_ok=True)

    for tidx, t in enumerate(data["transforms"]):
        db_ids = t.get("queryDbIds", [])
        for ridx, db_id in enumerate(db_ids):
            if db_id is None:
                continue
            if db_con:
                row = db_con.execute(
                    "SELECT query_text FROM queries WHERE id = ?", (db_id,)
                ).fetchone()
                text = row[0] if row else None
            else:
                text = None
            if not text:
                continue
            run_label  = data["runs"][ridx]["label"] if ridx < len(data.get("runs", [])) else f"run{ridx}"
            log_name   = f"{tidx+1:03d}__{_safe_filename(t['transform'])}__{_safe_filename(run_label)}.log"
            zip_name   = log_name + ".zip"
            zip_path   = os.path.join(logs_dir, zip_name)
            with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
                zf.writestr(log_name, text)
            # Store relative path so the template can trigger a plain <a download>
            if "queryLogFiles" not in t:
                t["queryLogFiles"] = [None] * len(db_ids)
            t["queryLogFiles"][ridx] = f"{logs_name}/{zip_name}"

    if db_con:
        db_con.close()

    with open(TEMPLATE_PATH, encoding="utf-8") as fh:
        template = fh.read()

    html = template.replace(DATA_MARKER, json.dumps(data, ensure_ascii=False))

    with open(output_path, "w", encoding="utf-8") as fh:
        fh.write(html)

    print(f"Dashboard written: {output_path}")
    print(f"Query log files:   {logs_dir}/")


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: generate_dashboard.py <input.xlsx> <output.html>")
        sys.exit(1)
    generate(sys.argv[1], sys.argv[2])
